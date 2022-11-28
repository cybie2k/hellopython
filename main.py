import pprint
from openpyxl import Workbook, load_workbook
import os
from datetime import timedelta,datetime
import logging
import pathlib
import pandas as pd
import numpy as np

logging.basicConfig(level=logging.DEBUG,filename='log.txt')
pp=pprint.PrettyPrinter(indent=4)

origcurdir=os.curdir
time_tracker_dir=os.path.join(os.getcwd(),'Time Tracker')
support_data_dir=os.path.join(time_tracker_dir,'Support Data')
support_data_file=os.path.join(support_data_dir,'Incident Hour Label.xlsx')
logging.debug(f"origcurdir={origcurdir}")
logging.debug(f"time_tracker_dir={time_tracker_dir}")
logging.debug(f"support_data_dir={support_data_dir}")
logging.debug(f"support_data_file={support_data_file}")

timelogdir=os.path.join(time_tracker_dir,'timelog')
logging.debug(f"timelogdir={timelogdir}")
timelogfiles=list(pathlib.Path(timelogdir).glob('*.xlsx'))
pth=str(timelogfiles[0])
logging.debug(pth)
logging.debug(f"timelogfiles={timelogfiles}")
##timelogfile=os.path.join(timelogdir,'TimeLog.xlsx')

support_dataframe=pd.read_excel(support_data_file)
#print(support_dataframe.to_string())


""" def dayofweekoffset(d):
    if (d<6) :
        return (d+1)*-1
    else:
        return 0 """

frames=[]
for timelogfile in timelogfiles:
    logging.debug(f"Reading file:{timelogfile}")
    tmpframe=pd.read_excel(timelogfile)
    frames.append(tmpframe[['Clock In','Clock Out','Task','Notes']])
    #frames.append(tmpframe)
TimeLogData=pd.concat(frames)


TimeLogData["Event Duration"]=(TimeLogData["Clock Out"]-TimeLogData["Clock In"])/pd.Timedelta(hours=1)
TimeLogData["Work Day"]=pd.to_datetime(TimeLogData['Clock In']-pd.Timedelta(hours=7)).dt.date
TimeLogData["Work Day of Week"]=pd.to_datetime(TimeLogData["Work Day"]).dt.day_of_week
TimeLogData["Temp"]=np.where(TimeLogData["Work Day of Week"]==6,0,(TimeLogData["Work Day of Week"]+1)*-1).astype(int)
TimeLogData["Work Week"]=(TimeLogData["Work Day"] - pd.to_datetime(TimeLogData["Work Day"]).dt.weekday * np.timedelta64(1,'D'))-pd.Timedelta(days=1)
##logging.debug(frames)

logging.debug(TimeLogData)








""" logging.info(f"timelogdir = {timelogdir}\ntimelogfile = {timelogfile}")
os.chdir(os.path.normpath(timelogdir))
logging.info(f"Change directory to {os.path.normpath(timelogdir)}")
wb = load_workbook(timelogfile,data_only=True)
ws = wb.worksheets[0]
logging.info(f"Loaded workbook {timelogfile}\nprocessing worksheet:{wb.worksheets[0]}")

for row in ws.iter_rows(min_row=1, max_row=1,values_only=True):
    rowheader=list(row)

pp.pprint(rowheader)

rownum=0
allrows={}

for row in ws.iter_rows(min_row=2,values_only=True):
    rowid=str(rownum)
    rowdata=list(row)
    zippedrow={rowheader:rowdata for rowheader,rowdata in zip(rowheader,rowdata)}
    ##allrows.append(zippedrow)
    edur=zippedrow['Clock Out']-zippedrow['Clock In']
    zippedrow['Event duration']=edur.seconds/3600
    allrows[rowid]=zippedrow
    ##print (rowid,'=',zippedrow)
  
    ##allrows{rownum}=zippedrow
    rownum+=1
##pp.pprint (allrows)
logging.info(f'Closing {timelogfile}')
wb.close
logging.info('Done!')
logging.warning('Danger! Danger! Will Robinson.') """